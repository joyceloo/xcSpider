# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook

class XcspiderPipeline(object):
	def __init__(self):
		self.wb=Workbook()
		self.ws=self.wb.active
		self.ws.append(['点评者','点评内容','出游时间','点评url','地点','省份'])
	def process_item(self, item, spider):
		line=[item['dp_user'][0],item['dp_content'][0],item['dp_time'][0],item['dp_link'][0],item['dp_scence'][0],item['dp_provice'][0]]
		self.ws.append(line)
		f=item['dp_scence'][0]
		self.wb.save(str(f)+'_dp.xlsx')
		return item

class YjspiderPipeline(object):
	def __init__(self):
		self.wb=Workbook()
		self.ws=self.wb.active
		self.ws.append(['标题','作者','浏览人数','评论人数','出游时间','游记url','省份'])
	def process_item(self, item, spider):
		line=[item['yj_title'][0],item['yj_author'][0],item['yj_looknum'][0],item['yj_pl'][0],item['yj_time'][0],item['yj_link'][0],item['yj_province'][0]]
		self.ws.append(line)
		f=item['yj_province'][0]
		self.wb.save(str(f)+'_yj.xlsx')
		return item

class AskspiderPipeline(object):
	def __init__(self):
		self.wb=Workbook()
		self.ws=self.wb.active
		self.ws.append(['问题','提问者','提问时间','链接','省份'])
	def process_item(self, item, spider):
		line=[item['q_title'][0],item['q_user'][0],item['q_time'][0],item['q_link'][0],item['q_province'][0]]
		self.ws.append(line)
		# f=item['q_province'][0]
		self.wb.save("北京"+'_wenda.xlsx')
		return item
